import os
import re
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import db, Employee, Attendance
import config

UPLOAD_EXTENSIONS = ['.xlsx', '.xls', '.csv']


def create_app():
    app = Flask(__name__)
    app.config.from_object('config')
    db.init_app(app)
    Migrate(app, db)

    os.makedirs(app.config.get('EXPORT_FOLDER', os.path.join(os.path.dirname(__file__), 'exports')), exist_ok=True)
    os.makedirs(app.config.get('UPLOAD_FOLDER', os.path.join(os.path.dirname(__file__), 'uploads')), exist_ok=True)

    with app.app_context():
        db.create_all()

    @app.route('/', methods=['GET', 'POST'])
    def index():
        if request.method == 'POST':
            f = request.files.get('file')
            if not f:
                flash('Aucun fichier téléchargé', 'warning')
                return redirect(url_for('index'))

            filename = secure_filename(f.filename)
            if not filename:
                flash('Nom de fichier invalide', 'warning')
                return redirect(url_for('index'))

            ext = os.path.splitext(filename)[1].lower()
            if ext not in UPLOAD_EXTENSIONS:
                flash('Type de fichier non supporté', 'danger')
                return redirect(url_for('index'))

            save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            f.save(save_path)

            try:
                rows_added = process_upload(save_path)
                flash(f'Fichier importé avec succès. Lignes ajoutées/mises à jour : {rows_added}', 'success')
            except Exception as e:
                flash(f'Erreur lors du traitement : {e}', 'danger')

            return redirect(url_for('index'))

        start = request.args.get('start')
        end = request.args.get('end')
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
            end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        except Exception:
            start_date = end_date = None

        recap_df = build_recap(start_date, end_date)
        table_html = recap_df.to_html(classes='table table-striped table-sm', index=False) if not recap_df.empty else '<p>Aucune donnée</p>'
        return render_template('index.html', table_html=table_html, start=start, end=end)

    @app.route('/export', methods=['GET'])
    def export():
        start = request.args.get('start')
        end = request.args.get('end')
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
            end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        except Exception:
            start_date = end_date = None

        if not start_date or not end_date:
            flash('Plage de dates invalide', 'warning')
            return redirect(url_for('index'))

        # Récupérer les données détaillées avec la plage de dates
        q = Attendance.query.join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )

        rows = q.with_entities(
            Attendance.date,
            Employee.matricule, Employee.nom, Employee.prenom, Employee.poste,
            Employee.site, Employee.affaire, Employee.classe, Employee.affectation, Employee.ville,
            Employee.taux_lgt, Employee.taux_repas,
            Attendance.present, Attendance.absent, Attendance.cong,
            Attendance.tour_rep, Attendance.repos_med, Attendance.sans_ph
        ).order_by(Employee.matricule.asc(), Attendance.date.asc()).all()

        # Récupérer tous les employés (même ceux sans données dans la plage)
        all_employees = Employee.query.order_by(Employee.matricule.asc()).all()

        if not all_employees:
            flash('Aucun employé trouvé', 'warning')
            return redirect(url_for('index'))

        # Créer toutes les dates de la plage
        dates_set = set()
        current_date = start_date
        while current_date <= end_date:
            dates_set.add(current_date)
            current_date += timedelta(days=1)
        
        sorted_dates = sorted(list(dates_set))

        # Préparer les données pour l'export
        employees_data = {}
        employee_totals = {}
        
        # Initialiser avec tous les employés
        for emp in all_employees:
            employees_data[emp.matricule] = {
                'matricule': emp.matricule,
                'nom': emp.nom or '',
                'prenom': emp.prenom or '',
                'poste': emp.poste or '',
                'site': emp.site or '',
                'affaire': emp.affaire or '',
                'classe': emp.classe or '',
                'affectation': emp.affectation or '',
                'ville': emp.ville or '',
                'taux_lgt': emp.taux_lgt or 0.0,
                'taux_repas': emp.taux_repas or 0.0,
                'attendances': {}
            }
            
            employee_totals[emp.matricule] = {
                'present': 0,
                'absent': 0,
                'cong': 0,
                'tour_rep': 0,
                'repos_med': 0,
                'sans_ph': 0
            }

        # Remplir avec les données existantes
        for r in rows:
            matricule = r[1]
            date_obj = r[0]
            
            if matricule not in employees_data:
                continue
                
            present_val = int(r[12] or 0)
            absent_val = int(r[13] or 0)
            cong_val = int(r[14] or 0)
            tour_rep_val = int(r[15] or 0)
            repos_med_val = int(r[16] or 0)
            sans_ph_val = int(r[17] or 0)
            
            employees_data[matricule]['attendances'][date_obj] = {
                'present': present_val,
                'absent': absent_val,
                'cong': cong_val,
                'tour_rep': tour_rep_val,
                'repos_med': repos_med_val,
                'sans_ph': sans_ph_val
            }
            
            # Mettre à jour les totaux
            employee_totals[matricule]['present'] += present_val
            employee_totals[matricule]['absent'] += absent_val
            employee_totals[matricule]['cong'] += cong_val
            employee_totals[matricule]['tour_rep'] += tour_rep_val
            employee_totals[matricule]['repos_med'] += repos_med_val
            employee_totals[matricule]['sans_ph'] += sans_ph_val

        # Construire les en-têtes
        fixed_columns = ['Matricule', 'Nom', 'Prénom', 'Poste', 'Site', 'Affaire', 'Classe', 'Affectation', 'Ville', 'Taux Logement', 'Taux Repas']
        
        # Couleurs pour les dates (cycle jaune, vert, bleu)
        date_colors = [
            PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # Jaune clair
            PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),  # Vert clair
            PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')   # Bleu clair
        ]
        
        # En-têtes pour les totaux
        total_headers_first = ['RÉCAPITULATIF'] * 6
        total_headers_second = ['Présent', 'Absent', 'CONG', 'Tour_rep', 'Repos_med', 'Sans_ph']
        
        # Créer les données pour le DataFrame
        columns_data = []
        
        # Préparer les données pour chaque employé
        for emp_data in employees_data.values():
            matricule = emp_data['matricule']
            row_data = [
                emp_data['matricule'],
                emp_data['nom'],
                emp_data['prenom'],
                emp_data['poste'],
                emp_data['site'],
                emp_data['affaire'],
                emp_data['classe'],
                emp_data['affectation'],
                emp_data['ville'],
                emp_data['taux_lgt'],
                emp_data['taux_repas']
            ]
            
            # Ajouter les statuts pour chaque date (même les dates sans données = 0)
            for date_obj in sorted_dates:
                attendance = emp_data['attendances'].get(date_obj, {})
                row_data.extend([
                    attendance.get('present', 0),
                    attendance.get('absent', 0),
                    attendance.get('cong', 0),
                    attendance.get('tour_rep', 0),
                    attendance.get('repos_med', 0),
                    attendance.get('sans_ph', 0)
                ])
            
            # Ajouter les totaux à la fin
            totals = employee_totals[matricule]
            row_data.extend([
                totals['present'],
                totals['absent'],
                totals['cong'],
                totals['tour_rep'],
                totals['repos_med'],
                totals['sans_ph']
            ])
            
            columns_data.append(row_data)
        
        # Construire les en-têtes complets
        first_header = fixed_columns.copy()
        second_header = [''] * len(fixed_columns)
        
        # Compter le nombre de colonnes pour chaque date (6 statuts)
        color_index = 0
        
        for date_obj in sorted_dates:
            date_str = date_obj.strftime('%d/%m/%Y')
            first_header.extend([date_str] * 6)
            second_header.extend(['Présent', 'Absent', 'CONG', 'Tour_rep', 'Repos_med', 'Sans_ph'])
            color_index += 1
        
        # Ajouter les en-têtes de récapitulatif
        first_header.extend(total_headers_first)
        second_header.extend(total_headers_second)
        
        # Chemin du fichier d'export avec plage de dates
        start_str = start.replace('-', '-') if start else datetime.now().strftime('%d-%m-%Y')
        end_str = end.replace('-', '-') if end else datetime.now().strftime('%d-%m-%Y')

        out_path = os.path.join(
            app.config['EXPORT_FOLDER'],
            f'presence_{start_str}_a_{end_str}.xlsx'
        )
        
        # Créer le fichier Excel manuellement avec openpyxl
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Feuille1'
        
        # Écrire la première ligne d'en-tête (dates fusionnées)
        col_num = 1
        # Colonnes fixes
        for header in fixed_columns:
            worksheet.cell(row=1, column=col_num, value=header)
            col_num += 1
        
        # Colonnes de dates fusionnées
        current_col = col_num
        date_color_map = {}
        
        for i, date_obj in enumerate(sorted_dates):
            date_str = date_obj.strftime('%d/%m/%Y')
            start_col = current_col
            end_col = current_col + 5  # 6 colonnes - 1 (index 0-based)
            
            # Fusionner les cellules pour la date
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            # Écrire la date au centre de la cellule fusionnée
            cell = worksheet.cell(row=1, column=start_col, value=date_str)
            
            # Stocker la position et la couleur pour cette date
            date_color_map[(start_col, end_col)] = date_colors[i % len(date_colors)]
            
            current_col = end_col + 1
        
        # Colonnes de récapitulatif fusionnées
        recap_start_col = current_col
        recap_end_col = current_col + 5
        worksheet.merge_cells(start_row=1, start_column=recap_start_col, end_row=1, end_column=recap_end_col)
        worksheet.cell(row=1, column=recap_start_col, value='RÉCAPITULATIF')
        
        # Écrire la deuxième ligne d'en-tête (statuts)
        col_num = 1
        # Colonnes fixes (vide pour la deuxième ligne)
        for header in fixed_columns:
            worksheet.cell(row=2, column=col_num, value='')
            col_num += 1
        
        # Statuts pour chaque date
        for date_obj in sorted_dates:
            statuts = ['Présent', 'Absent', 'CONG', 'Tour_rep', 'Repos_med', 'Sans_ph']
            for statut in statuts:
                worksheet.cell(row=2, column=col_num, value=statut)
                col_num += 1
        
        # Statuts pour le récapitulatif
        statuts_recap = ['Présent', 'Absent', 'CONG', 'Tour_rep', 'Repos_med', 'Sans_ph']
        for statut in statuts_recap:
            worksheet.cell(row=2, column=col_num, value=statut)
            col_num += 1
        
        # Écrire les données
        for row_num, row_data in enumerate(columns_data, 3):  # Commencer à la ligne 3
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)
        
        # Appliquer le style aux en-têtes
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Appliquer le style à la première ligne d'en-tête
        for col_num in range(1, len(first_header) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Appliquer le style à la deuxième ligne d'en-tête
        for col_num in range(1, len(second_header) + 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Appliquer les couleurs aux dates fusionnées
        for (start_col, end_col), color_fill in date_color_map.items():
            for col_num in range(start_col, end_col + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = color_fill
                # Appliquer aussi la couleur aux en-têtes de statuts de cette date
                statut_cell = worksheet.cell(row=2, column=col_num)
                statut_cell.fill = color_fill
        
        # Appliquer une couleur différente pour le récapitulatif
        recap_color = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')  # Orange clair
        for col_num in range(recap_start_col, recap_end_col + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = recap_color
            statut_cell = worksheet.cell(row=2, column=col_num)
            statut_cell.fill = recap_color
        
        # Ajuster la largeur des colonnes automatiquement
        for col_num in range(1, len(first_header) + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row_num in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier
        workbook.save(out_path)

        return send_file(out_path, as_attachment=True, download_name=os.path.basename(out_path))

    return app


# -------------------------
# Fonctions utilitaires
# -------------------------

def _normalize(s: str) -> str:
    """Normalise une chaîne pour comparaison d'en-têtes."""
    if s is None:
        return ''
    return re.sub(r'[^A-Za-z0-9éèêàôùïçÉÀÈ]', '', str(s)).strip().lower()


def _is_date_string(s: str) -> bool:
    """Detecte jj/mm/yyyy (ou j/m/yyyy)."""
    if not s:
        return False
    s = str(s).strip()
    return bool(re.match(r'^\d{1,2}\/\d{1,2}\/\d{4}$', s))


def _parse_date_flexible(date_str: str):
    """Parse une date en essayant plusieurs formats."""
    date_str = str(date_str).strip()
    
    # Formats à essayer
    formats = [
        '%d/%m/%Y',  # DD/MM/YYYY
        '%m/%d/%Y',  # MM/DD/YYYY
        '%Y-%m-%d',  # YYYY-MM-DD
        '%d-%m-%Y',  # DD-MM-YYYY
        '%m-%d-%Y',  # MM-DD-YYYY
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    # Si aucun format ne fonctionne, essayer pandas
    try:
        parsed = pd.to_datetime(date_str, dayfirst=False, errors='coerce')
        if not pd.isna(parsed):
            return parsed.date()
    except:
        pass
    
    return None


def process_upload(path: str) -> int:
    """
    Lit un fichier Excel avec double en-tête (header=[0,1]) où la ligne 1 contient
    les dates (fusionnées en Excel) et la ligne 2 contient les statuts (Présent/Absent/...).
    Insère / met à jour les attendances.
    Retourne le nombre d'enregistrements d'attendance traités.
    """
    # Lecture en header multi-index
    try:
        df = pd.read_excel(path, header=[0, 1], engine='openpyxl')
    except Exception as e:
        print(f"DEBUG: Error reading with multi-index: {e}")
        # fallback : lecture simple
        df = pd.read_excel(path, engine='openpyxl')

    # Convertir tuples en listes manipulables
    cols = list(df.columns)

    # normaliser tous les éléments en str et forward-fill le niveau "top" si vide
    top_level = []
    bottom_level = []
    last_top = None
    for a, b in cols:
        a_s = str(a).strip() if a is not None else ''
        b_s = str(b).strip() if b is not None else ''
        # consider pandas "Unnamed" or empty as blank
        if a_s == '' or a_s.lower().startswith('unnamed') or a_s.lower() == 'nan':
            a_s = last_top or ''
        else:
            last_top = a_s
        top_level.append(a_s)
        bottom_level.append(b_s)

    # rebuild MultiIndex labels as tuples (top, bottom)
    new_cols = [(top_level[i], bottom_level[i]) for i in range(len(top_level))]
    df.columns = pd.MultiIndex.from_tuples(new_cols)

    # canonical fixed columns and variants
    canonical_fixed = {
        'matricule': ['matricule', 'id'],
        'nom': ['nom', 'name'],
        'prenom': ['prenom', 'prénom', 'prenom'],
        'poste': ['poste', 'position'],
        'site': ['site'],
        'affaire': ['affaire'],
        'classe': ['classe', 'class', 'niveau'],
        'affectation': ['affectation', 'affect', 'assignment'],
        'ville': ['ville', 'city'],
        'taux_logement': ['tauxlogement', 'taux_logement', 'taux_lgt', 'taux logement', 'tauxlgt'],
        'taux_repas': ['tauxrepas', 'taux_repas', 'taux repas', 'tauxrep']
    }

    # find fixed columns (they may be in top or bottom depending on how template was made)
    fixed_map = {}
    for col in df.columns:
        top, bot = col
        top_n = _normalize(top)
        bot_n = _normalize(bot)
        for canon, variants in canonical_fixed.items():
            if top_n in variants or bot_n in variants:
                fixed_map[canon] = col
                break

    # Build date_columns: mapping date_str -> {status_label: column_tuple}
    date_columns = {}
    for col in df.columns:
        top, bot = col
        # Essayer de parser la date avec plusieurs formats
        date_candidate = None
        for candidate in (top, bot):
            if candidate:
                parsed_date = _parse_date_flexible(str(candidate))
                if parsed_date:
                    date_candidate = parsed_date
                    # L'autre élément est le statut
                    status_candidate = bot if candidate == top else top
                    break
        
        if date_candidate:
            date_str = date_candidate.strftime('%Y-%m-%d')  # Stocker en format standard
            status = str(status_candidate).strip() if status_candidate else ''
            date_columns.setdefault(date_str, {})[status] = col

    print(f"DEBUG: Found {len(date_columns)} date columns in import file: {list(date_columns.keys())}")

    rows_processed = 0

    # DB transaction: we'll commit at end (explicitly)
    try:
        for idx, row in df.iterrows():
            # 1. Identification de l'employé par Matricule
            matricule = None
            if 'matricule' in fixed_map:
                matricule = row[fixed_map['matricule']]
            else:
                # Heuristique de secours: prendre la première colonne
                matricule = row[df.columns[0]]

            if pd.isna(matricule):
                continue
            matricule = str(matricule).strip()
            if not matricule:
                continue

            # find or create employee
            emp = Employee.query.filter_by(matricule=matricule).first()
            if not emp:
                emp = Employee(matricule=matricule)
                db.session.add(emp)
                db.session.flush()

            # 2. Mise à jour des informations fixes de l'employé
            for key in ('nom', 'prenom', 'poste', 'site', 'affaire', 'classe', 'affectation', 'ville'):
                if key in fixed_map:
                    val = row[fixed_map[key]]
                    if pd.notna(val):
                        try:
                            setattr(emp, key, str(val).strip())
                        except Exception:
                            pass

            # Mise à jour des taux (logement / repas)
            if 'taux_logement' in fixed_map:
                try:
                    val = row[fixed_map['taux_logement']]
                    if pd.notna(val):
                        emp.taux_lgt = float(val)
                except Exception:
                    # ignore bad conversion
                    pass
            if 'taux_repas' in fixed_map:
                try:
                    val = row[fixed_map['taux_repas']]
                    if pd.notna(val):
                        emp.taux_repas = float(val)
                except Exception:
                    pass

            # 3. Parcours des dates et statuts (le cœur de l'import)
            for date_str, status_map in date_columns.items():
                try:
                    # date_str est déjà au format YYYY-MM-DD
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                except Exception:
                    # if parse fails skip this date
                    continue

                # status flags
                present_flag = 0
                absent_flag = 0
                cong_flag = 0
                tourrep_flag = 0
                repos_flag = 0
                sansph_flag = 0

                def cell_true(val):
                    if pd.isna(val):
                        return False
                    if isinstance(val, (int, float)):
                        return val != 0
                    s = str(val).strip().lower()
                    if not s:
                        return False
                    return s in ('x', '1', 'yes', 'y', 'présent', 'present', 'p') or (s.isdigit() and int(s) != 0)

                # read each status cell for the date
                for status_label, colkey in status_map.items():
                    try:
                        sval = row[colkey]
                        if cell_true(sval):
                            sn = str(status_label).strip().lower()
                            if 'prés' in sn or 'present' in sn:
                                present_flag = 1
                            elif 'abs' in sn:
                                absent_flag = 1
                            elif 'cong' in sn:
                                cong_flag = 1
                            elif 'tour' in sn and 'rep' in sn:
                                tourrep_flag = 1
                            elif 'repos' in sn or 'méd' in sn or 'med' in sn:
                                repos_flag = 1
                            elif 'sans' in sn and 'ph' in sn:
                                sansph_flag = 1
                    except Exception:
                        pass

                # skip if no status flagged
                if (present_flag + absent_flag + cong_flag + tourrep_flag + repos_flag + sansph_flag) == 0:
                    continue

                # upsert attendance row (employee_id + date)
                att = Attendance.query.filter_by(employee_id=emp.id, date=date_obj).first()
                if not att:
                    att = Attendance(
                        employee_id=emp.id,
                        date=date_obj,
                        present=present_flag,
                        absent=absent_flag,
                        cong=cong_flag,
                        tour_rep=tourrep_flag,
                        repos_med=repos_flag,
                        sans_ph=sansph_flag
                    )
                    db.session.add(att)
                else:
                    # replace with new values (user imports explicit statuses)
                    att.present = present_flag
                    att.absent = absent_flag
                    att.cong = cong_flag
                    att.tour_rep = tourrep_flag
                    att.repos_med = repos_flag
                    att.sans_ph = sansph_flag

                rows_processed += 1

        db.session.commit()
        print(f"DEBUG: Successfully processed {rows_processed} attendance records")
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Erreur lors du traitement des données: {e}")

    return rows_processed


def build_recap(start_date=None, end_date=None):
    """
    Retourne un DataFrame récapitulatif par employé (totaux) pour la plage fournie (inclusive).
    Colonnes : Matricule, Nom, Prénom, Poste, Site, Affaire, Classe, Affectation, Ville,
              Présent, Absent, CONG, Tour_rep, Repos_med, Sans_ph
    """
    q = Attendance.query.join(Employee)

    if start_date:
        q = q.filter(Attendance.date >= start_date)
    if end_date:
        q = q.filter(Attendance.date <= end_date)

    # Aggregation par employé
    q = q.with_entities(
        Employee.matricule,
        Employee.nom,
        Employee.prenom,
        Employee.poste,
        Employee.site,
        Employee.affaire,
        Employee.classe,
        Employee.affectation,
        Employee.ville,
        db.func.coalesce(db.func.sum(Attendance.present), 0).label('Présent'),
        db.func.coalesce(db.func.sum(Attendance.absent), 0).label('Absent'),
        db.func.coalesce(db.func.sum(Attendance.cong), 0).label('CONG'),
        db.func.coalesce(db.func.sum(Attendance.tour_rep), 0).label('Tour_rep'),
        db.func.coalesce(db.func.sum(Attendance.repos_med), 0).label('Repos_med'),
        db.func.coalesce(db.func.sum(Attendance.sans_ph), 0).label('Sans_ph')
    ).group_by(Employee.id, Employee.matricule, Employee.nom, Employee.prenom, Employee.poste,
               Employee.site, Employee.affaire, Employee.classe, Employee.affectation, Employee.ville
    ).order_by(Employee.matricule.asc())

    rows = q.all()
    if not rows:
        return pd.DataFrame()

    data = []
    for r in rows:
        data.append({
            'Matricule': r[0],
            'Nom': r[1] or '',
            'Prénom': r[2] or '',
            'Poste': r[3] or '',
            'Site': r[4] or '',
            'Affaire': r[5] or '',
            'Classe': r[6] or '',
            'Affectation': r[7] or '',
            'Ville': r[8] or '',
            'Présent': int(r[9] or 0),
            'Absent': int(r[10] or 0),
            'CONG': int(r[11] or 0),
            'Tour_rep': int(r[12] or 0),
            'Repos_med': int(r[13] or 0),
            'Sans_ph': int(r[14] or 0)
        })

    df = pd.DataFrame(data)
    return df


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True)